import os, re, pytesseract, shutil
from PIL import Image
from datetime import datetime
from openpyxl import Workbook, load_workbook
from telegram.ext import Updater, MessageHandler, Filters

# 📊 Kayıt dosyaları
EXCEL_PATH = "XMuhasebe.xlsx"
IMAGE_DIR = "Fisler"
os.makedirs(IMAGE_DIR, exist_ok=True)

# 📘 Eğer Excel dosyası yoksa oluştur
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.title = "Yönetici"
    ws.append(["ID","Yükleme Tarihi","Belge Tarihi","Firma","Tutar","Döviz","Kullanıcı","Fotoğraf Yolu"])
    wb.save(EXCEL_PATH)

# 🔍 OCR işlemi (fotoğraftan metin okuma)
def ocr_parse(image_path):
    text = pytesseract.image_to_string(Image.open(image_path), lang='tur')

    # Belge tarihi (örnek: 30.10.2025)
    date_match = re.search(r'(\d{2}[./-]\d{2}[./-]\d{4})', text)
    doc_date = date_match.group(0) if date_match else ""

    # Tutar (örnek: 120,00)
    amt_match = re.search(r'(?<!\d)(\d{1,3}(?:[.,]\d{3})*[.,]\d{2})(?!\d)', text)
    amount = amt_match.group(1) if amt_match else ""

    # Firma adı
    firm_match = re.search(r'([A-ZÇĞİÖŞÜ]{2,}(?:\s+[A-ZÇĞİÖŞÜ&.-]{2,}){0,3})', text)
    firm = firm_match.group(0).title() if firm_match else ""

    return text, doc_date, amount, firm

# 💾 Excel'e kayıt ekleme
def save_excel(user_name, upload_dt, doc_dt, firm, amount, image_path, currency="TRY"):
    wb = load_workbook(EXCEL_PATH)
    if user_name not in wb.sheetnames:
        ws = wb.create_sheet(user_name)
        ws.append(["ID","Yükleme Tarihi","Belge Tarihi","Firma","Tutar","Döviz","Kullanıcı","Fotoğraf Yolu"])
    ws = wb[user_name]
    admin_ws = wb["Yönetici"]
    row_id = ws.max_row
    ws.append([row_id, upload_dt, doc_dt, firm, amount, currency, user_name, image_path])
    admin_ws.append([row_id, upload_dt, doc_dt, firm, amount, currency, user_name, image_path])
    wb.save(EXCEL_PATH)

# 🤖 Telegram bot fotoğraf işlemi
def handle_photo(update, context):
    user_name = update.message.from_user.first_name or "Bilinmeyen"
    file_obj = update.message.photo[-1].get_file()
    tmp_path = f"/tmp/tmp_{datetime.now().timestamp()}.jpg"
    file_obj.download(tmp_path)

    text, doc_date, amount, firm = ocr_parse(tmp_path)

    # Fotoğrafı kaydet
    file_name = f"FIS_{datetime.now().strftime('%Y-%m-%d_%H-%M')}_{user_name}.jpg"
    drive_path = os.path.join(IMAGE_DIR, file_name)
    shutil.copy(tmp_path, drive_path)

    # Excel'e ekle
    save_excel(user_name, datetime.now().strftime("%Y-%m-%d %H:%M"), doc_date, firm, amount, drive_path)
    update.message.reply_text(
        f"✅ Kayıt eklendi {user_name}.\n"
        f"🏢 Firma: {firm or '-'}\n"
        f"📅 Belge Tarihi: {doc_date or '-'}\n"
        f"💰 Tutar: {amount or '-'}\n"
        f"📸 Fotoğraf: Kaydedildi."
    )
    os.remove(tmp_path)

# 🔑 Token (Render'da environment değişkeninden alır)
TOKEN = os.getenv("BOT_TOKEN")

updater = Updater(TOKEN, use_context=True)
dp = updater.dispatcher
dp.add_handler(MessageHandler(Filters.photo, handle_photo))

print("🤖 XMuhasebe Bot 7/24 aktif (Render).")
updater.start_polling()
updater.idle()
